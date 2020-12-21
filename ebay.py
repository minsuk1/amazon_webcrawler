#-*- coding: utf-8 -*-

import requests
import json
import re
import time
import datetime
import sys
import os
from openpyxl import Workbook
import random
from openpyxl import load_workbook
from random import randint
from urllib.parse import quote
from bs4 import BeautifulSoup
import html
from selenium import webdriver
from selenium.webdriver.common.keys import Keys

now = time.localtime()
s = "%04d%02d%02d%02d%02d%02d" % (now.tm_year, now.tm_mon, now.tm_mday, now.tm_hour, now.tm_min, now.tm_sec)
s1 = "%04d%02d%02d" % (now.tm_year, now.tm_mon, now.tm_mday)

delay_list = []
f = open("delay.ini", 'r')
while True:
    line = f.readline()
    if not line: break
    delay_list.append(re.sub('\n',"",line))
f.close()

page_list = []
f = open("page.ini", 'r')
while True:
    line = f.readline()
    if not line: break
    page_list.append(re.sub('\n',"",line))
f.close()

replypage_list = []
f = open("reply_page.ini", 'r')
while True:
    line = f.readline()
    if not line: break
    replypage_list.append(re.sub('\n',"",line))
f.close()

def parser_list(all, start, end, content):
    content = content.replace("\n", "")
    content = content.replace("\r", "")
    content = content.replace("\t", "")
    partten = re.compile(all)
    p_start = re.compile(start)
    p_end = re.compile(end)
    partten_temp = []
    result_list = []
    partten_temp = partten.findall(content)
    for value in partten_temp:
        value = re.sub(start, '', value)
        value = re.sub(end, '', value)
        #print (value)
        result_list.append(value)        
    return result_list

def parser_str(all, start, end, content):
    content = content.replace("\n", "")
    content = content.replace("\r", "")
    content = content.replace("\t", "")
    partten = re.compile(all)
    p_start = re.compile(start)
    p_end = re.compile(end)
    partten_temp = []
    result_str = ''
    partten_temp = partten.findall(content)
    for value in partten_temp:
        value = re.sub(start, '', value)
        value = re.sub(end, '', value)
        #print (value)
        result_str = value
        break        
    return result_str

headers_common = {'Host':'www.ebay.com','Connection':'close','Cache-Control':'max-age=0','Upgrade-Insecure-Requests':'1','User-Agent':'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/80.0.3987.163 Safari/537.36','Sec-Fetch-Dest':'document','Accept':'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9','Sec-Fetch-Site':'same-origin','Sec-Fetch-Mode':'navigate','Sec-Fetch-User':'?1','Accept-Encoding':'gzip, deflate','Accept-Language':'ko,ko-KR;q=0.9,en-US;q=0.8,en;q=0.7,ja;q=0.6','Content-Length':'4'}

wb_save = Workbook()
ws_save = wb_save.active                
title_cnt=1

titles  = ['No','URL','Code','content','상품명']
for title in titles:
    ws_save.cell(row=1, column=title_cnt).value = title
    title_cnt = title_cnt + 1

wb_reply = Workbook()
ws_reply = wb_reply.active                
title_cnt=1

titles  = ['No','Code','상품명','날짜','리뷰제목','리뷰내용','별점','상태']
for title in titles:
    ws_reply.cell(row=1, column=title_cnt).value = title
    title_cnt = title_cnt + 1

wb = load_workbook('category.xlsx')
sheetList = wb.get_sheet_names()
sheet = wb.get_sheet_by_name(sheetList[0])

key_cnt = 2
try:
    while True:
        category = sheet.cell(row=key_cnt, column=1).value
        if category is None:
            break
        page_cnt = 1
        cnt = 2
        reply_cnt = 2
        print (category)

        response = requests.get(category,headers=headers_common)
        category_str = ''
        category_str = parser_str('<span class=\"b-pageheader\_\_text\">.*?<','<span class=\"b-pageheader\_\_text\">','<',response.text)
        category_str = re.sub('/', ',', category_str)
        print (category_str)
        if not os.path.isdir('./'+str(category_str)+'/'):
            os.mkdir('./'+str(category_str)+'/')
        else:
            pass
        for page_cnt in range(int(page_list[0]),int(page_list[1])+1):

            print ("===keyword "+str(key_cnt-1)+" page "+str(page_cnt)+"===")
            url = category+'&_pgn='+str(page_cnt)
            response = requests.get(url,headers=headers_common)
            prod_list = []
            prod_list = parser_list('<a class=\"s-item\_\_link\" href=\".*?\"','<a class=\"s-item\_\_link\" href=\"','\"',response.text)
            prod_title = parser_list('class=\"s-item\_\_title\".*?>.*?<','class=\"s-item\_\_title\".*?>','<',response.text)
            print (len(prod_list))
            print (prod_list)
            if len(prod_list) == 0:
                break
            prod_cnt = 0
            for prod_url in prod_list:
                url = prod_url
                print (url)
                url = 'https://www.ebay.com/p/182457545?iid=233149353460'
                response = requests.get(url,headers=headers_common)
                iid = parser_str(';iid=.*?\"',';iid=','\"',response.text)
                print (iid)
                epid = parser_str('epid=.*?\"','epid=','\"',response.text)
                if epid == '':
                    epid = parser_str('\"epid\":\".*?\"','\"epid\":\"','\"',response.text)
                #print (BeautifulSoup(brand, "html.parser").text.strip())

                ws_save.cell(row=cnt, column=1).value = cnt-1
                ws_save.cell(row=cnt, column=2).value = url
                ws_save.cell(row=cnt, column=3).value = iid
                ws_save.cell(row=cnt, column=5).value = prod_title[prod_cnt]
                prod_cnt = prod_cnt + 1

                title = parser_str('og:title\" Content=\".*?\"','og:title\" Content=\"','\"',response.text)
                title = title.replace('&quot;','"')
                title = title.replace('| eBay','')
                title = title.replace('for sale online','')

                product_details = parser_str('<div id=\"ProductDetails\".*?</section>','','',response.text)                
                if product_details == '':
                    product_details = parser_str('<div class=\"app-itemspecifics.*?chevron\"></div></div></div>','','',response.text)
                    product_details = product_details.replace('</span>',' ')
                    if product_details == '':
                        product_details = parser_str('<h2 class=\"secHd\">Item specifics</h2>.*?</table>','','',response.text)
                #print (BeautifulSoup(product_details, "html.parser").text.strip())    
                ws_save.cell(row=cnt, column=4).value = BeautifulSoup(product_details, "html.parser").text.strip()


                try:
                    second_num = random.randint(int(delay_list[0]),int(delay_list[1]))
                except:
                    second_num = delay_list[0]
                print ("[Delay] wait "+str(second_num)+" second")
                time.sleep(int(second_num))

                cnt = cnt + 1

                if not os.path.isdir('./'+str(category_str)+'/'):
                    os.mkdir('./'+str(category_str)+'/')
                else:
                    pass
        
                wb_save.save('./'+str(category_str)+'/product_'+s+'_result.xlsx')

                for page_cnt in range(int(replypage_list[0]),int(replypage_list[1])+1):
                    print ("===reply page "+str(page_cnt)+"===")
                    url_review = "https://www.ebay.com/urw/product-reviews/"+str(epid)
                    url = url_review+'?pgn='+str(page_cnt)
                    print (url)
                    response = requests.get(url,headers=headers_common)
                    reply_data = parser_list('ebay-review-section\".*?</div></div></div>','','',response.text)
                    if len(reply_data) == 0:
                        break
                    for reply in reply_data:
                        ws_reply.cell(row=reply_cnt, column=1).value = reply_cnt-1
                        ws_reply.cell(row=reply_cnt, column=2).value = iid
                        ws_reply.cell(row=reply_cnt, column=3).value = title
                        
                        reply_date = parser_str('class=\\\"review-item-date\\\">.*?</span>','class=\\\"review-item-date\\\">','</span>',reply)
                        ws_reply.cell(row=reply_cnt, column=4).value = reply_date
                        print (reply_date)
                        
                        reply_title = parser_str('class=\\\"review-item-title wrap-spaces\\\">.*?</h3>','class=\\\"review-item-title wrap-spaces\\\">','</h3>',reply)
                        ws_reply.cell(row=reply_cnt, column=5).value = reply_title
                        
                        reply_content = parser_str('<p itemprop=\"reviewBody.*?</p>','','',reply)
                        ws_reply.cell(row=reply_cnt, column=6).value = BeautifulSoup(reply_content, "html.parser").text.strip()
                        
                        reply_star = parser_str('\"><meta itemprop=\"ratingValue\" content=\".*?\"','\"><meta itemprop=\"ratingValue\" content=\"','\\\"',reply)
                        ws_reply.cell(row=reply_cnt, column=7).value = reply_star
                        
                        condition =  parser_str('capitalize\">.*?<','capitalize\">','<',reply)
                        ws_reply.cell(row=reply_cnt, column=8).value = condition

                        reply_cnt = reply_cnt + 1

                    wb_reply.save('./'+str(category_str)+'/review_'+s+'_result.xlsx')
                    
                    try:
                        second_num = random.randint(int(delay_list[0]),int(delay_list[1]))
                    except:
                        second_num = delay_list[0]
                    print ("[Delay] wait "+str(second_num)+" second")
                    time.sleep(int(second_num))
                

                       
            page_cnt = page_cnt + 1            
 
        key_cnt = key_cnt +1
     
except Exception as err:
        print ('Error on line {}'.format(sys.exc_info()[-1].tb_lineno))
        print ("%s.\n" % str(err))
finally:
    input('--- program end ---')


